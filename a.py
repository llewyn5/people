import openpyxl

wb = openpyxl.load_workbook('aa.xlsx')

# print(wb.sheetnames)
#
# for sheet in wb:
#     print(sheet.title)
# mySheet = wb.create_chartsheet("mySheet")
# print(wb.sheetnames)

#sheet3 = wb.get_sheet_by_name('sheet3')
#sheet4 = wb['mysheet']
#保存之后才时改写了文件
wb.save('aa.xlsx')

#获取活跃的表单
ws = wb.active
# print(ws)
# print(ws['A1'])
# print(ws['A1'].value)
#
# c = ws['B1']
# print('Row {},Colume {}is{}'.format(c.row,c.column,c.value))
# print('Cell {} is {}'.format(c.coordinate,c.value))
#
# print(ws.cell(row= 1,column=2))
# print(ws.cell(row= 1,column=2).value)
#
# for i in range(1,8,2):
#     print(i,ws.cell(row=i,column=2).value)

#从表单中取行和列
# colC = ws['C']
# print(colC[0].value)
# col_range = ws['B:C']
# row_range = ws[2:6]

# for col in col_range:
#     for cell in col:
#         print(cell.value)
#
# for row in row_range:
#     for cell in row:
#         print(cell.value)


for row in ws.iter_rows(min_row=1,max_row=2,max_col=2):
    for cell in row:
        print(cell)

cell_range = ws['A1:C3']
for a in cell_range:
    for b in a:
        print(b.coordinate,b.value)
    print('-------------------')


print('{} * {}'.format(ws.max_column,ws.max_row))

from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(2),get_column_letter(247))
print(column_index_from_string('HH'))